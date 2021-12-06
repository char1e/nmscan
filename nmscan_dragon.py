import sys
import os
import openpyxl
import nmap  
import time
from concurrent.futures import ThreadPoolExecutor,ProcessPoolExecutor

"""
输入文件格式为ip:port
与nmscan相比，端口默认已经扫描出来了，不使用masscan
version:2
每次扫描出结果产生一个新工作表

"""
def myNmap(nmapInfo):

    host = nmapInfo['host']     
    arguments = nmapInfo['arguments']
    scan = nmap.PortScanner()       
    
    print('nmap正在扫描host:',end='')
    print(nmapInfo['host'])
    print('使用参数:',end='')
    print(nmapInfo['arguments'])
    
    scan_result = scan.scan(hosts=host,arguments=arguments)     #创建一次扫描，并行执行，取出结果
    print("执行命令:" + scan.command_line())    #不知道为啥，这个选项就是会返回多余的-oX - 例如u'nmap -oX - -p 22,80 -sV 192.168.209.121-122'
    
    print('scan_result:')
    print(scan_result,end='\n\n')

    tcpInfo = {}
    #hostname = scan_result['scan'][host]['hostnames'][0]['name'] #主机名,可能有很多主机名此处取第一个
    address = scan_result['scan'][host]['addresses']['ipv4']    #主机ip地址
    status = scan_result['scan'][host]['status']['state']   #主机状态:up或者down

    tcpInfo = {}
    udpInfo = {}
    ports_count = 0
    tcp_ports = []
    udp_ports = []
    ip_ports = []
    sctp_ports = []
    all_protocols = scan[host].all_protocols()
    for protocol in all_protocols:
        tcp_ports = scan[host].all_tcp() #所有tcp端口列表
        udp_ports = scan[host].all_udp() #
        ip_ports = scan[host].all_ip() #
        sctp_ports = scan[host].all_sctp() #
    ports_count = len(tcp_ports) + len(udp_ports) + len(ip_ports) + len(sctp_ports)
    
    if(ports_count > len(tcp_ports)):
        print('发现除TCP外别的端口' + '\n\n')
    else:
        print('只有TCP端口')
    
    print('%s 主机端口数量为 %d' % (host,ports_count))
    if ports_count == 0:
        print('%s 无端口，跳过该IP' % (host))      #masscan发现端口，但是nmap扫描的时候又没有发现此端口
        return 
        
        
    if ports_count > 1000:
        print("%s端口太多可能有waf",host)
    
    print(len(tcp_ports))
    print(len(udp_ports))
    
    #写入TCP信息

    print('line132')
    if len(tcp_ports) > 0:
        for tcp_port in tcp_ports:
            tcp_port_info = scan[host]['tcp'][tcp_port]
            
            #tcpInfo[host][tcp_port] = {}
            #print('tcp_port_info:')
            #print(tcp_port_info)       #字典类型带大括号
           
            tcpInfo[tcp_port] = tcp_port_info
            #info[host]["ports_nmap"]["ports"] = info[host]["ports_nmap"]["ports"].update(tcp_port,tcp_port_info)
        #针对每一个host
        print('%s的TCP端口信息为'%(host),end='\n')
        print(tcpInfo,end='\n\n')
        nmapData2Excel(host,tcpInfo,'Tcp')
    
    #写入UDP信息
    if len(udp_ports) > 0:
        for udp_port in udp_ports:
            udp_port_info = scan[host]['udp'][udp_port]
            udpInfo[udp_port] = udp_port_info
        print('%s的UDP端口信息为'%(host),end='\n')
        print(udpInfo,end='\n\n')
        nmapData2Excel(host,udpInfo,'Udp')
            



def nmapData2Excel(host,hostInfo,protocol): #写入xls文件，如果存在就修改
    print('\n进入excel表格处理nmapData2Excel函数,处理IP:%s\n' % (host))
    
    ports = list(hostInfo.keys())
    print('端口列表:',end='')
    print(ports)
    
    if not os.path.exists("./nmscanOutput/"+host+".xlsx"):
        #创建一个新工作表
        workbook = openpyxl.Workbook()
        workSheet = workbook.active
        workSheet.title = time.strftime("%Y-%m-%d %H时%M分%S秒",time.localtime())
    else:
        #工作表已存在
        workbook = openpyxl.load_workbook("./nmscanOutput/"+host+".xlsx")
        workSheet = workbook.create_sheet(time.strftime("%Y-%m-%d %H时%M分%S秒",time.localtime()))
    
    print('创建名为:',end='')
    print(time.strftime("%Y-%m-%d %H时%M分%S秒",time.localtime()),end='')
    print('的工作表')
    
    #写列名
    columns = ["ip","port","state","name","product","version","cpe","extrainfo",'script']
    #columns = ['ip','port'] + list(hostInfo.[ports[0]].keys())
    lenCol = len(columns)
    for j in range(1,lenCol+1):
            workSheet.cell(1,j,columns[j-1])
    #workSheet.cell(1,j+1,'script')
    
    
    #写数据
    i = 2
    for port in ports:
        workSheet.cell(i,1,host)     
        workSheet.cell(i,2,port)
        #写IP和port之后的列
        for k in range(2,lenCol-1):
            workSheet.cell(i,k+1,hostInfo[port][columns[k]])
        if 'script' in hostInfo[port]:
            scriptContent = str(hostInfo[port][columns[lenCol-1]]).strip('{').strip('}')
            workSheet.cell(i,lenCol,scriptContent)
        i=i+1

    workbook.save(filename="./nmscanOutput/"+ host + ".xlsx")
    print(host + '的%s工作表操作完成' % (time.strftime("%Y-%m-%d %H时%M分%S秒",time.localtime()),))
    print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),end='\n\n')
    
    
    
if __name__ == "__main__":
    targetFile = sys.argv[1]
    while True:
        if os.path.exists(targetFile):
            break
        else:
            print('Invalid targetFile\n')
            targetsFile = input = ("please input targetFile")
            
    ipList = []
    portList = []
    file = open(targetFile)
    while True:
        text =  file.readline()
        if text:
            tmpList = text.split(':')
            ipList.append(tmpList[0])
            portList.append(tmpList[1])
        else:
            break
    #print(IPList)
    nmapParam = ' -sV -Pn -sS -v --script="http-title"'
    nmapThreads = 20
    thread_pool = ThreadPoolExecutor(nmapThreads)
    
    if not os.path.exists("./nmscanOutput"):
        os.makedirs("./nmscanOutput")
    
    for i in range(len(ipList)):
        nmapFuncParam = {}
        nmapFuncParam['host'] = ipList[i]
        nmapPortsParam = ' -p ' + portList[i].strip('\n') #nmap只扫描masscan发现的端口，拼接成-p xxx的选项
        nmapFuncParam['arguments'] = nmapParam + nmapPortsParam      #最终nmap执行的选项
        print(nmapFuncParam)
        
        thread_pool.submit(myNmap,nmapFuncParam)

   
   
