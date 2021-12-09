import openpyxl
import time
import os
import json
def nmapData2Excel(host,hostInfo,protocol): #写入xls文件，如果存在就修改
    print('\n进入excel表格处理nmapData2Excel函数,处理IP:%s\n' % (host))
    
    ports = list(hostInfo.keys())
    print('端口列表:',end='')
    print(ports)
    
    if not os.path.exists("./nmscanOutput/"+host+".xlsx"):
        workbook = openpyxl.Workbook()
    else:
        workbook = openpyxl.load_workbook("./nmscanOutput/"+host+".xlsx")
    
    if protocol+'Ports' in workbook.sheetnames:
        workSheet = workbook[protocol+'Ports']
    else:
        workSheet = workbook.create_sheet(protocol + 'Ports',0)
        workSheet = workbook[protocol + 'Ports']
    
    #写列名
    columns = ["ip","port","state","name","product","version","cpe","extrainfo",'script']
    #columns = ['ip','port'] + list(hostInfo.[ports[0]].keys())
    lenCol = len(columns)
    for j in range(1,lenCol+1):
            workSheet.cell(1,j,columns[j-1])
    #workSheet.cell(1,j+1,'script')
    
    
    #写数据
    i = 2
    for port in ports:
        workSheet.cell(i,1,host)     
        workSheet.cell(i,2,port)
        #写IP和port之后的列
        for k in range(2,lenCol-1):
            workSheet.cell(i,k+1,hostInfo[port][columns[k]])
        if 'script' in hostInfo[port]:
            scriptContent = str(hostInfo[port][columns[lenCol-1]]).strip('{').strip('}')
            workSheet.cell(i,lenCol,scriptContent)
        i=i+1

    workbook.save(filename="./nmscanOutput/"+ host + ".xlsx")
    print(host + '的excel表格（%s页）操作完成' % (protocol))
    print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),end='\n\n')
