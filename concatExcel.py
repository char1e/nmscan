#本脚本用于将nmscan生成的多个xlsx文件内容合并到同一个文件当中
#使用方法：python3 concatExcel.py <nmscanOutput文件夹路径>
#会自动在同目录下生成concatResult.xlsx的结果文件
#by rhaps

import openpyxl
import os
import sys
from openpyxl import load_workbook,Workbook


targetPath = sys.argv[1]
#创建工作表
wbResult = Workbook()
wsResult = wbResult.active


#如果为True则增加ip:port列
flag = False;
flag = True;
if flag:
    columns = ["ip","port","ip:port","state","name","product","version","cpe","extrainfo",'script']
else:
    columns = ["ip","port","state","name","product","version","cpe","extrainfo",'script']

lenCol = len(columns)
for j in range(1,lenCol+1):
        wsResult.cell(1,j,columns[j-1])
wsResultRowCount = 2;

for filePath, dirNames, fileNames in os.walk(targetPath):
    for fileName in fileNames: #fileName是单纯的文件名，filePathName是完整的文件路径
        filePathName = os.path.join(os.getcwd(),filePath,fileName)
        print("Processing: " + str(filePathName))
        wb = openpyxl.load_workbook(filePathName)
        
        for ws in wb:
            maxRow = ws.max_row
            maxCol = ws.max_column
            
            if flag:
                for i in range(2,maxRow+1):
                    for j in range(1, maxCol+2):
                        if(j < 3):
                            value = ws.cell(i,j).value
                            wsResult.cell(wsResultRowCount, j, value)
                        elif(j == 3):
                            value = str(ws.cell(i,j-2).value) + ":" + str(ws.cell(i,j-1).value)
                            wsResult.cell(wsResultRowCount, j, value)
                        else:
                            value = ws.cell(i,j-1).value
                            wsResult.cell(wsResultRowCount, j, value)
                    wsResultRowCount = wsResultRowCount + 1
            else:
                for i in range(2,maxRow+1):
                    for j in range(1, maxCol+1):
                        value = ws.cell(i,j).value
                        wsResult.cell(wsResultRowCount, j, value)
                    wsResultRowCount = wsResultRowCount + 1
        print(60)
        wb.close();
print(62)
wbResult.save(filePath + '.xlsx')